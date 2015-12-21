'''
Created on 21 de oct. de 2015

@author: cgimenop
@updated Oct 29th 2015

'''


import getopt
import re
import sys

from TLSuite import TLSuite
from TLCase import TLCase

sys.path.insert(1,"lib/openpyxl")
sys.path.insert(1,"lib/jdcal-1.0")
sys.path.insert(1,"lib/et_xmlfile")

from xmlfile import *
from jdcal import *
from openpyxl import load_workbook

class InvalidParameterException(Exception):
    def __init__(self, name, value):
        self.parameter = name
        self.value = value
    
    def __str__(self):
        return " ".join(["Invalid parameter", self.parameter], "with value", self.value)

def read_excel_test_plan(filename="", test_plan="" , suite_name="", start_row=-1, end_row=-1):
    full_test_plan = load_workbook(filename, read_only=True)
    # TODO: Put all this into a class to reuse and allow call by the main script sheet by sheet parametrized
    # Does not work 4 unknown reason
    # ws = test_plan['big_data'] # ws is now an IterableWorksheet
    
    # This will be an user parameter
    tests = full_test_plan.get_sheet_by_name(test_plan)
    
            
    # Beware some E2E test will be missing with this pattern => E2E_1.x.2
    test_id_pattern = re.compile("^[A-Z0-9]+_[A-Z0-9]+.[A-Z0-9]+.[A-Z0-9]+$", re.IGNORECASE)
    # story_pattern = re.compile("^Story:.*$", re.IGNORECASE)

    # service_suites = None;
    # endpoint_suites = None;
    method_suite = TLSuite(suite_name);
    # TODO: Somehow must detect which rows have content ideally automatic, user parameter may ve an option
    # for row in tests.iter_rows('A3:N32'):
    
    for row in  tests.iter_rows('A'+str(start_row)+':N'+str(end_row)):
        
        if (row[0].value is not None):
            if (test_id_pattern.match(row[0].value) and method_suite is not None):      
                case = TLCase(row)
                method_suite.add_tlcase(case)
                
    return method_suite

def write_testlink_xml(test_suite, filename=""):
    if (filename is ""):
        raise
    xml_str = test_suite.to_xml()
    
    with open(filename, "w") as f:
        f.write(xml_str)
                
        

''' MAIN '''
try:
    opts, args = getopt.getopt(sys.argv[1:], "", ["filename=", "sheet=", "start=", "end=", "suite=", "output="])
except getopt.GetoptError as err:
    print ("migrateExcelTestSuite.py --filename <excel_test_plan> --sheet <test_plan_sheet> --start <start_row> --end <end_row> --suite <ouput_test_suite_name> --output <output_filename>")
    print(str(err))
    sys.exit(2)

input_file = ""
sheet = ""
suite = ""
output_file = ""
start_row = -1
end_row = -1

for opt, value in opts:
    if (opt == "--filename"):
        input_file = value
    elif (opt == "--sheet"):
        sheet = value     
    elif (opt == "--suite"):
        suite = value
    elif (opt == "--output"):
        output_file = value
    elif (opt == "--start"):
        start_row = int(value)
    elif (opt == "--end"):
        end_row = int(value)
    else:
        print ("Unsupported option "+opt +" with value "+ value)
        print ("migrateExcelTestSuite.py --filename <excel_test_plan> --sheet <test_plan_sheet> --start <start_row> --end <end_row> --suite <ouput_test_suite_name> --output <output_filename>")
        sys.exit(2)
        
#FIXME: Upper limit for excel row is not checked. Should be added
#FIXME: No control oover start_row > end_row
if (input_file == "" or sheet == "" or suite == "" or output_file == "" or start_row <= 0 or end_row <= 0):
    print ("One or more mandatory parameters are empty or have a invalid value")
    print ("migrateExcelTestSuite.py --filename <excel_test_plan> --sheet <test_plan_sheet> --start <start_row> --end <end_row> --suite <ouput_test_suite_name> --output <output_filename>")
    sys.exit(2)


write_testlink_xml(read_excel_test_plan(input_file,sheet, suite, start_row, end_row), output_file)

     
        









